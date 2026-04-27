# -*- coding: utf-8 -*-
"""
backend/services/sentiment_semantic_service.py - 第五期情感与语义分析服务

提供情感分析和语义聚类的数据查询与任务触发功能。
复用 backend/database.py 中已有的数据库连接方法。
"""

import os
import subprocess
import sys
from pathlib import Path
from typing import Optional

import pandas as pd
import pymysql
from pymysql.cursors import DictCursor

# 获取项目根目录
PROJECT_ROOT = Path(__file__).parent.parent.parent.resolve()

# 添加项目根目录到 sys.path
sys.path.insert(0, str(PROJECT_ROOT))


def _get_connection():
    """
    获取 MySQL 数据库连接。

    Returns:
        pymysql.connections.Connection: 数据库连接对象
    """
    from backend.config import settings

    return pymysql.connect(
        host=settings.mysql_host,
        port=settings.mysql_port,
        user=settings.mysql_user,
        password=settings.mysql_password,
        database=settings.mysql_database,
        charset=settings.mysql_charset,
        cursorclass=DictCursor,
        autocommit=True,
    )


def _serialize_datetime(obj):
    """
    序列化 datetime/date 对象为字符串。

    Args:
        obj: datetime 或 date 对象

    Returns:
        str or original: 如果是 datetime/date 则返回字符串，否则返回原对象
    """
    if hasattr(obj, "strftime"):
        return obj.strftime("%Y-%m-%d %H:%M:%S")
    return obj


def _query_to_dict_list(cursor):
    """
    将查询结果转换为字典列表，并序列化 datetime 字段。

    Args:
        cursor: pymysql 游标对象

    Returns:
        list: 字典列表
    """
    results = cursor.fetchall()
    for row in results:
        for key, value in row.items():
            if isinstance(value, (pd.Timestamp,)):
                row[key] = _serialize_datetime(value)
            elif isinstance(value, object) and hasattr(value, "strftime"):
                row[key] = _serialize_datetime(value)
    return results


def get_sentiment_summary():
    """
    获取情感分析总体概览。

    查询 hot_search_sentiment_stats，返回：
    - avg_sentiment_score: 平均情感分数
    - positive_count: 正向数量
    - neutral_count: 中性数量
    - negative_count: 负向数量
    - total_count: 总数量

    Returns:
        dict: 情感概览数据
    """
    conn = None
    try:
        conn = _get_connection()
        cursor = conn.cursor()

        sql = """
            SELECT
                AVG(sentiment_score) as avg_sentiment_score,
                SUM(CASE WHEN sentiment_label = 'positive' THEN 1 ELSE 0 END) as positive_count,
                SUM(CASE WHEN sentiment_label = 'neutral' THEN 1 ELSE 0 END) as neutral_count,
                SUM(CASE WHEN sentiment_label = 'negative' THEN 1 ELSE 0 END) as negative_count,
                COUNT(*) as total_count
            FROM hot_search_sentiment_stats
        """
        cursor.execute(sql)
        result = cursor.fetchone()

        if result is None or not result.get("total_count"):
            print("[report] hot_search_sentiment_stats 为空，情感汇总无数据")
            return {
                "avg_sentiment_score": 0.5,
                "positive_count": 0,
                "neutral_count": 0,
                "negative_count": 0,
                "total_count": 0,
            }

        # 处理数值
        for key in ["positive_count", "neutral_count", "negative_count", "total_count"]:
            if result.get(key) is not None:
                result[key] = int(result[key])
            else:
                result[key] = 0

        if result.get("avg_sentiment_score") is not None:
            result["avg_sentiment_score"] = round(float(result["avg_sentiment_score"]), 4)
        else:
            result["avg_sentiment_score"] = 0.5

        return result

    except pymysql.MySQLError as e:
        if "doesn't exist" in str(e):
            raise Exception("情感分析结果表不存在，请先执行 sql/v5_sentiment_semantic.sql 创建表")
        raise Exception(f"查询情感分析概览失败：{e}")
    finally:
        if conn:
            conn.close()


def get_sentiment_daily():
    """
    获取每日情感统计数据。

    查询 hot_search_sentiment_daily_stats，按日期升序排序。

    Returns:
        list: 每日情感统计数据列表
    """
    conn = None
    try:
        conn = _get_connection()
        cursor = conn.cursor()

        sql = """
            SELECT
                stat_date,
                avg_sentiment_score,
                positive_count,
                neutral_count,
                negative_count,
                total_count,
                positive_ratio,
                neutral_ratio,
                negative_ratio
            FROM hot_search_sentiment_daily_stats
            ORDER BY stat_date ASC
        """
        cursor.execute(sql)
        results = _query_to_dict_list(cursor)
        if not results:
            print("[report] hot_search_sentiment_daily_stats 为空，每日情绪无数据")
        return results

    except pymysql.MySQLError as e:
        if "doesn't exist" in str(e):
            raise Exception("每日情感统计表不存在，请先执行 sql/v5_sentiment_semantic.sql 创建表")
        raise Exception(f"查询每日情感统计失败：{e}")
    finally:
        if conn:
            conn.close()


def get_sentiment_top(limit: int = 20, label: Optional[str] = None):
    """
    获取情感分析 Top 数据。

    Args:
        limit: 返回条数限制，范围 1-100
        label: 可选，按情感标签过滤（positive/neutral/negative）

    Returns:
        list: 情感分析结果列表
    """
    if limit < 1:
        limit = 1
    elif limit > 100:
        limit = 100

    conn = None
    try:
        conn = _get_connection()
        cursor = conn.cursor()

        if label and label in ("positive", "neutral", "negative"):
            sql = f"""
                SELECT
                    keyword,
                    sentiment_score,
                    sentiment_label,
                    sentiment_label_cn,
                    rank_num,
                    hot_value,
                    fetch_time
                FROM hot_search_sentiment_stats
                WHERE sentiment_label = %s
                ORDER BY hot_value DESC
                LIMIT %s
            """
            cursor.execute(sql, (label, limit))
        else:
            sql = """
                SELECT
                    keyword,
                    sentiment_score,
                    sentiment_label,
                    sentiment_label_cn,
                    rank_num,
                    hot_value,
                    fetch_time
                FROM hot_search_sentiment_stats
                ORDER BY sentiment_score DESC, hot_value DESC
                LIMIT %s
            """
            cursor.execute(sql, (limit,))

        results = _query_to_dict_list(cursor)
        if not results:
            print("[report] hot_search_sentiment_stats 为空，情感明细无数据")
        return results

    except pymysql.MySQLError as e:
        if "doesn't exist" in str(e):
            raise Exception("情感分析结果表不存在，请先执行 sql/v5_sentiment_semantic.sql 创建表")
        raise Exception(f"查询情感分析 Top 数据失败：{e}")
    finally:
        if conn:
            conn.close()


def search_sentiment(keyword: str):
    """
    根据关键词模糊查询情感分析结果。

    Args:
        keyword: 搜索关键词

    Returns:
        list: 匹配的情感分析结果列表
    """
    conn = None
    try:
        conn = _get_connection()
        cursor = conn.cursor()

        sql = """
            SELECT
                keyword,
                sentiment_score,
                sentiment_label,
                sentiment_label_cn,
                rank_num,
                hot_value,
                fetch_time
            FROM hot_search_sentiment_stats
            WHERE keyword LIKE %s
            ORDER BY hot_value DESC
            LIMIT 100
        """
        cursor.execute(sql, (f"%{keyword}%",))
        results = _query_to_dict_list(cursor)
        return results

    except pymysql.MySQLError as e:
        if "doesn't exist" in str(e):
            raise Exception("情感分析结果表不存在，请先执行 sql/v5_sentiment_semantic.sql 创建表")
        raise Exception(f"搜索情感分析结果失败：{e}")
    finally:
        if conn:
            conn.close()


def get_semantic_clusters():
    """
    获取语义聚类结果。

    Returns:
        list: 语义聚类结果列表，按 cluster_id 和 hot_value 排序
    """
    conn = None
    try:
        conn = _get_connection()
        cursor = conn.cursor()

        sql = """
            SELECT
                keyword,
                cluster_id,
                cluster_name,
                semantic_keywords,
                embedding_method,
                hot_value,
                rank_num,
                cluster_date,
                created_at
            FROM hot_search_semantic_clusters
            ORDER BY cluster_id ASC, hot_value DESC
        """
        cursor.execute(sql)
        results = _query_to_dict_list(cursor)
        if not results:
            print("[report] 语义聚类结果为空，请先运行 python semantic/semantic_cluster_job.py")
        return results

    except pymysql.MySQLError as e:
        if "doesn't exist" in str(e):
            raise Exception("语义聚类结果表不存在，请先执行 sql/v5_sentiment_semantic.sql 创建表")
        raise Exception(f"查询语义聚类结果失败：{e}")
    finally:
        if conn:
            conn.close()


def get_semantic_cluster_summary():
    """
    获取语义聚类主题分布统计。

    Returns:
        list: 按 cluster_name 分组的统计数据
    """
    conn = None
    try:
        conn = _get_connection()
        cursor = conn.cursor()

        sql = """
            SELECT
                cluster_name,
                COUNT(*) as count,
                AVG(hot_value) as avg_hot_value,
                MAX(hot_value) as max_hot_value
            FROM hot_search_semantic_clusters
            GROUP BY cluster_name
            ORDER BY count DESC
        """
        cursor.execute(sql)
        results = _query_to_dict_list(cursor)
        if not results:
            print("[report] 语义聚类主题分布为空，请先运行 python semantic/semantic_cluster_job.py")

        # 处理数值
        for row in results:
            for key in ["count", "avg_hot_value", "max_hot_value"]:
                if row.get(key) is not None:
                    if key == "count":
                        row[key] = int(row[key])
                    else:
                        row[key] = round(float(row[key]), 2)

        return results

    except pymysql.MySQLError as e:
        if "doesn't exist" in str(e):
            raise Exception("语义聚类结果表不存在，请先执行 sql/v5_sentiment_semantic.sql 创建表")
        raise Exception(f"查询语义聚类主题分布失败：{e}")
    finally:
        if conn:
            conn.close()


def run_sentiment_job():
    """
    触发情感分析任务。

    使用 subprocess 调用 python sentiment/sentiment_job.py

    Returns:
        dict: 执行结果
    """
    try:
        script_path = PROJECT_ROOT / "sentiment" / "sentiment_job.py"
        result = subprocess.run(
            [sys.executable, str(script_path)],
            capture_output=True,
            text=True,
            timeout=300,  # 5分钟超时
        )

        return {
            "success": result.returncode == 0,
            "message": result.stdout if result.returncode == 0 else result.stderr,
            "returncode": result.returncode,
        }

    except subprocess.TimeoutExpired:
        raise Exception("情感分析任务执行超时（5分钟）")
    except Exception as e:
        raise Exception(f"执行情感分析任务失败：{str(e)}")


def run_semantic_cluster_job():
    """
    触发语义聚类任务。

    使用 subprocess 调用 python semantic/semantic_cluster_job.py

    Returns:
        dict: 执行结果
    """
    try:
        script_path = PROJECT_ROOT / "semantic" / "semantic_cluster_job.py"
        result = subprocess.run(
            [sys.executable, str(script_path)],
            capture_output=True,
            text=True,
            timeout=600,  # 10分钟超时
        )

        return {
            "success": result.returncode == 0,
            "message": result.stdout if result.returncode == 0 else result.stderr,
            "returncode": result.returncode,
        }

    except subprocess.TimeoutExpired:
        raise Exception("语义聚类任务执行超时（10分钟）")
    except Exception as e:
        raise Exception(f"执行语义聚类任务失败：{str(e)}")


def _get_keyword_stats():
    """获取关键词统计（用于报告）"""
    conn = None
    try:
        conn = _get_connection()
        cursor = conn.cursor()

        # 使用 best_rank AS min_rank_num 确保报告字段名一致
        sql = """
            SELECT
                keyword,
                appear_count,
                max_hot_value,
                best_rank AS min_rank_num
            FROM hot_search_keyword_stats
            ORDER BY appear_count DESC, max_hot_value DESC
            LIMIT 100
        """
        cursor.execute(sql)
        results = _query_to_dict_list(cursor)
        if not results:
            print("[report] hot_search_keyword_stats 为空，关键词统计无数据")
        return results
    except Exception as e:
        print(f"[report] 获取关键词统计失败：{e}")
        return []
    finally:
        if conn:
            conn.close()


def _get_daily_stats():
    """获取每日统计（用于报告）"""
    conn = None
    try:
        conn = _get_connection()
        cursor = conn.cursor()

        # 真实表字段：total_records / total_keywords / min_rank（无 created_at / keyword_count / total_count）
        sql = """
            SELECT
                stat_date,
                total_records AS total_count,
                total_keywords AS keyword_count,
                avg_hot_value,
                max_hot_value,
                min_rank
            FROM hot_search_daily_stats
            ORDER BY stat_date ASC
        """
        cursor.execute(sql)
        results = _query_to_dict_list(cursor)
        if not results:
            print("[report] hot_search_daily_stats 为空，每日统计无数据")
        return results
    except Exception as e:
        print(f"[report] 获取每日统计失败：{e}")
        return []
    finally:
        if conn:
            conn.close()


def _get_burst_predictions():
    """获取爆发趋势预测（用于报告）"""
    conn = None
    try:
        conn = _get_connection()
        cursor = conn.cursor()

        # 使用 SQL CASE 表达式在数据库层生成 burst_label，避免 Python 端循环
        sql = """
            SELECT
                keyword,
                burst_probability,
                burst_level,
                CASE burst_level
                    WHEN 2 THEN '爆发型'
                    WHEN 1 THEN '稳定型'
                    WHEN 0 THEN '降温型'
                    ELSE '稳定型'
                END AS burst_label,
                trend_direction,
                current_hot_value
            FROM hot_search_burst_predictions
            ORDER BY burst_probability DESC, current_hot_value DESC
            LIMIT 100
        """
        cursor.execute(sql)
        results = _query_to_dict_list(cursor)
        if not results:
            print("[report] hot_search_burst_predictions 为空，爆发趋势无数据")
        return results
    except Exception as e:
        print(f"[report] 获取爆发趋势失败：{e}")
        return []
    finally:
        if conn:
            conn.close()


def _get_topic_clusters():
    """获取主题聚类（用于报告）"""
    conn = None
    try:
        conn = _get_connection()
        cursor = conn.cursor()

        sql = """
            SELECT
                keyword,
                cluster_id,
                cluster_name,
                tfidf_keywords,
                hot_value,
                rank_num,
                cluster_date,
                created_at
            FROM hot_search_topic_clusters
            ORDER BY cluster_id ASC, hot_value DESC
            LIMIT 200
        """
        cursor.execute(sql)
        results = _query_to_dict_list(cursor)
        if not results:
            print("[report] hot_search_topic_clusters 为空，TF-IDF 主题聚类无数据")
        return results
    except Exception as e:
        print(f"[report] 获取主题聚类失败：{e}")
        return []
    finally:
        if conn:
            conn.close()


def _get_current_ranking():
    """获取当前热搜榜（用于报告）"""
    conn = None
    try:
        conn = _get_connection()
        cursor = conn.cursor()

        sql = """
            SELECT
                title,
                rank_num,
                hot_value,
                source,
                fetch_time
            FROM hot_search_raw
            WHERE fetch_time = (SELECT MAX(fetch_time) FROM hot_search_raw)
            ORDER BY rank_num ASC
            LIMIT 50
        """
        cursor.execute(sql)
        results = _query_to_dict_list(cursor)
        if not results:
            print("[report] hot_search_raw 为空，当前热搜榜无数据")
        return results
    except Exception as e:
        print(f"[report] 获取当前热搜榜失败：{e}")
        return []
    finally:
        if conn:
            conn.close()


def _get_burst_label(burst_level):
    """根据 burst_level 生成中文爆发标签。"""
    try:
        level = int(burst_level)
    except (TypeError, ValueError):
        level = 0
    if level == 3:
        return "高爆发"
    if level == 2:
        return "中爆发"
    return "低爆发"


def get_report_debug_counts():
    """获取增强报告相关关键表的数据量。"""
    tables = [
        "hot_search_raw",
        "hot_search_keyword_stats",
        "hot_search_daily_stats",
        "hot_search_feature_stats",
        "hot_search_burst_predictions",
        "hot_search_topic_clusters",
        "hot_search_sentiment_stats",
        "hot_search_sentiment_daily_stats",
        "hot_search_semantic_clusters",
    ]
    counts = {}
    conn = None
    try:
        conn = _get_connection()
        cursor = conn.cursor()
        for table in tables:
            try:
                cursor.execute(f"SELECT COUNT(*) AS count FROM {table}")
                row = cursor.fetchone() or {}
                counts[table] = int(row.get("count") or 0)
                if counts[table] == 0:
                    if table == "hot_search_semantic_clusters":
                        print("[report] 语义聚类结果为空，请先运行 python semantic/semantic_cluster_job.py")
                    else:
                        print(f"[report] {table} 为空")
            except Exception as e:
                counts[table] = None
                print(f"[report] 获取 {table} 数据量失败：{e}")
        return counts
    except Exception as e:
        print(f"[report] 获取调试表数据量失败：{e}")
        return {table: None for table in tables}
    finally:
        if conn:
            conn.close()


def export_enhanced_report_csv():
    """
    导出增强版综合报告 CSV。

    包含：情感分析汇总、每日情绪统计、语义聚类结果、
    第四期爆发趋势识别结果、第三期关键词统计结果、当前热搜榜

    Returns:
        tuple: (csv_content, filename)
    """
    import io

    try:
        # 获取所有数据
        sentiment_summary = get_sentiment_summary()
        sentiment_daily = get_sentiment_daily()
        sentiment_top = get_sentiment_top(limit=100)
        semantic_clusters = get_semantic_clusters()
        keyword_stats = _get_keyword_stats()
        daily_stats = _get_daily_stats()
        burst_predictions = _get_burst_predictions()
        topic_clusters = _get_topic_clusters()
        current_ranking = _get_current_ranking()

        # 构建 CSV
        output = io.StringIO()

        # 情感分析汇总
        output.write("=== 情感分析汇总 ===\n")
        output.write(f"平均情感分数,{sentiment_summary.get('avg_sentiment_score', 0)}\n")
        output.write(f"正向数量,{sentiment_summary.get('positive_count', 0)}\n")
        output.write(f"中性数量,{sentiment_summary.get('neutral_count', 0)}\n")
        output.write(f"负向数量,{sentiment_summary.get('negative_count', 0)}\n")
        output.write(f"总数量,{sentiment_summary.get('total_count', 0)}\n\n")

        # 每日情绪统计
        output.write("=== 每日情绪统计 ===\n")
        output.write("日期,平均情感分数,正向数量,中性数量,负向数量,总数量,正向占比,中性占比,负向占比\n")
        for row in sentiment_daily:
            output.write(f"{row.get('stat_date', '')},{row.get('avg_sentiment_score', 0)},"
                        f"{row.get('positive_count', 0)},{row.get('neutral_count', 0)},{row.get('negative_count', 0)},"
                        f"{row.get('total_count', 0)},{row.get('positive_ratio', 0)},{row.get('neutral_ratio', 0)},{row.get('negative_ratio', 0)}\n")
        output.write("\n")

        # 每日热搜统计
        output.write("=== 每日热搜统计 ===\n")
        output.write("日期,总记录数,关键词数,平均热度,最高热度,最佳排名,创建时间\n")
        for row in daily_stats:
            output.write(f"{row.get('stat_date', '')},{row.get('total_records', 0)},"
                        f"{row.get('total_keywords', 0)},{row.get('avg_hot_value', 0)},"
                        f"{row.get('max_hot_value', 0)},{row.get('min_rank', '')},"
                        f"{row.get('created_at', '')}\n")
        output.write("\n")

        # 情感分析 Top100
        output.write("=== 情感分析 Top100 ===\n")
        output.write("关键词,情感分数,情感标签,排名,热度值\n")
        for row in sentiment_top:
            output.write(f"{row.get('keyword', '')},{row.get('sentiment_score', 0)},"
                        f"{row.get('sentiment_label_cn', '')},{row.get('rank_num', '')},{row.get('hot_value', 0)}\n")
        output.write("\n")

        # 语义聚类结果
        output.write("=== 语义聚类结果 ===\n")
        output.write("关键词,聚类ID,主题类别,语义关键词,嵌入方法,热度值,排名\n")
        for row in semantic_clusters:
            output.write(f"{row.get('keyword', '')},{row.get('cluster_id', '')},{row.get('cluster_name', '')},"
                        f"{row.get('semantic_keywords', '')},{row.get('embedding_method', '')},"
                        f"{row.get('hot_value', 0)},{row.get('rank_num', '')}\n")
        output.write("\n")

        # 爆发趋势识别
        output.write("=== 爆发趋势识别 Top100 ===\n")
        output.write("关键词,爆发等级,爆发类型,爆发概率,趋势方向,当前排名,当前热度值,热度变化率,排名变化,出现次数,模型名称,预测日期,创建时间\n")
        for row in burst_predictions:
            output.write(f"{row.get('keyword', '')},{row.get('burst_level', '')},"
                        f"{row.get('burst_label', '')},{row.get('burst_probability', 0)},"
                        f"{row.get('trend_direction', '')},{row.get('current_rank', '')},"
                        f"{row.get('current_hot_value', 0)},{row.get('hot_value_change_rate', 0)},"
                        f"{row.get('rank_change', 0)},{row.get('appear_count', 0)},"
                        f"{row.get('model_name', '')},{row.get('predict_date', '')},{row.get('created_at', '')}\n")
        output.write("\n")

        # 主题聚类
        output.write("=== TF-IDF主题聚类 ===\n")
        output.write("关键词,聚类ID,主题类别,TF-IDF关键词,热度值,排名,聚类日期,创建时间\n")
        for row in topic_clusters:
            output.write(f"{row.get('keyword', '')},{row.get('cluster_id', '')},{row.get('cluster_name', '')},"
                        f"{row.get('tfidf_keywords', '')},{row.get('hot_value', 0)},"
                        f"{row.get('rank_num', '')},{row.get('cluster_date', '')},{row.get('created_at', '')}\n")
        output.write("\n")

        # 关键词统计
        output.write("=== 关键词统计 Top100 ===\n")
        output.write("关键词,出现次数,最高热度,平均热度,最佳排名,最新采集时间,统计日期\n")
        for row in keyword_stats:
            output.write(f"{row.get('keyword', '')},{row.get('appear_count', 0)},"
                        f"{row.get('max_hot_value', 0)},{row.get('avg_hot_value', 0)},"
                        f"{row.get('best_rank', '')},{row.get('latest_fetch_time', '')},"
                        f"{row.get('stat_date', '')}\n")
        output.write("\n")

        # 当前热搜榜
        output.write("=== 当前热搜榜 ===\n")
        output.write("标题,排名,热度值,来源,采集时间\n")
        for row in current_ranking:
            output.write(f"{row.get('title', '')},{row.get('rank_num', '')},"
                        f"{row.get('hot_value', 0)},{row.get('source', '')},{row.get('fetch_time', '')}\n")

        csv_content = output.getvalue()
        output.close()

        return csv_content, "weibo_hot_enhanced_report.csv"

    except Exception as e:
        raise Exception(f"导出增强报告 CSV 失败：{str(e)}")


def export_enhanced_report_excel():
    """
    导出增强版综合报告 Excel（多 Sheet）。

    Returns:
        tuple: (excel_file_path, filename)
    """
    from datetime import datetime
    import tempfile

    try:
        import openpyxl
        from openpyxl import Workbook

        # 创建工作簿
        wb = Workbook()

        # 移除默认 sheet
        wb.remove(wb.active)

        # 获取所有数据
        sentiment_summary = get_sentiment_summary()
        sentiment_daily = get_sentiment_daily()
        sentiment_top = get_sentiment_top(limit=100)
        semantic_clusters = get_semantic_clusters()
        semantic_summary = get_semantic_cluster_summary()
        keyword_stats = _get_keyword_stats()
        daily_stats = _get_daily_stats()
        burst_predictions = _get_burst_predictions()
        topic_clusters = _get_topic_clusters()
        current_ranking = _get_current_ranking()

        # Sheet1: 情感汇总
        ws1 = wb.create_sheet("情感汇总")
        ws1.append(["微博热搜情感分析汇总"])
        ws1.append(["平均情感分数", sentiment_summary.get("avg_sentiment_score", 0)])
        ws1.append(["正向数量", sentiment_summary.get("positive_count", 0)])
        ws1.append(["中性数量", sentiment_summary.get("neutral_count", 0)])
        ws1.append(["负向数量", sentiment_summary.get("negative_count", 0)])
        ws1.append(["总数量", sentiment_summary.get("total_count", 0)])

        # Sheet2: 每日情绪
        ws2 = wb.create_sheet("每日情绪")
        ws2.append(["日期", "平均情感分数", "正向数量", "中性数量", "负向数量",
                     "总数量", "正向占比", "中性占比", "负向占比"])
        for row in sentiment_daily:
            ws2.append([
                row.get("stat_date", ""),
                row.get("avg_sentiment_score", 0),
                row.get("positive_count", 0),
                row.get("neutral_count", 0),
                row.get("negative_count", 0),
                row.get("total_count", 0),
                row.get("positive_ratio", 0),
                row.get("neutral_ratio", 0),
                row.get("negative_ratio", 0),
            ])

        # Sheet: 每日统计
        ws_daily = wb.create_sheet("每日统计")
        ws_daily.append(["日期", "总记录数", "关键词数", "平均热度", "最高热度", "最佳排名", "创建时间"])
        for row in daily_stats:
            ws_daily.append([
                row.get("stat_date", ""),
                row.get("total_records", 0),
                row.get("total_keywords", 0),
                row.get("avg_hot_value", 0),
                row.get("max_hot_value", 0),
                row.get("min_rank", ""),
                row.get("created_at", ""),
            ])

        # Sheet3: 情感明细
        ws3 = wb.create_sheet("情感明细")
        ws3.append(["关键词", "情感分数", "情感标签", "排名", "热度值"])
        for row in sentiment_top:
            ws3.append([
                row.get("keyword", ""),
                row.get("sentiment_score", 0),
                row.get("sentiment_label_cn", ""),
                row.get("rank_num", ""),
                row.get("hot_value", 0),
            ])

        # Sheet4: 语义聚类
        ws4 = wb.create_sheet("语义聚类")
        ws4.append(["关键词", "聚类ID", "主题类别", "语义关键词", "嵌入方法", "热度值", "排名"])
        for row in semantic_clusters:
            ws4.append([
                row.get("keyword", ""),
                row.get("cluster_id", ""),
                row.get("cluster_name", ""),
                row.get("semantic_keywords", ""),
                row.get("embedding_method", ""),
                row.get("hot_value", 0),
                row.get("rank_num", ""),
            ])

        # Sheet5: 语义主题分布
        ws5 = wb.create_sheet("语义主题分布")
        ws5.append(["主题类别", "数量", "平均热度", "最高热度"])
        for row in semantic_summary:
            ws5.append([
                row.get("cluster_name", ""),
                row.get("count", 0),
                row.get("avg_hot_value", 0),
                row.get("max_hot_value", 0),
            ])

        # Sheet: TF-IDF主题聚类
        ws_topic = wb.create_sheet("TF-IDF主题聚类")
        ws_topic.append(["关键词", "聚类ID", "主题类别", "TF-IDF关键词", "热度值", "排名", "聚类日期", "创建时间"])
        for row in topic_clusters:
            ws_topic.append([
                row.get("keyword", ""),
                row.get("cluster_id", ""),
                row.get("cluster_name", ""),
                row.get("tfidf_keywords", ""),
                row.get("hot_value", 0),
                row.get("rank_num", ""),
                row.get("cluster_date", ""),
                row.get("created_at", ""),
            ])

        # Sheet6: 爆发趋势
        ws6 = wb.create_sheet("爆发趋势")
        ws6.append(["关键词", "爆发等级", "爆发类型", "爆发概率", "趋势方向", "当前排名", "当前热度值", "热度变化率", "排名变化"])
        for row in burst_predictions:
            ws6.append([
                row.get("keyword", ""),
                row.get("burst_level", ""),
                row.get("burst_label", ""),
                row.get("burst_probability", 0),
                row.get("trend_direction", ""),
                row.get("current_rank", ""),
                row.get("current_hot_value", 0),
                row.get("hot_value_change_rate", 0),
                row.get("rank_change", 0),
            ])

        # Sheet7: 关键词统计
        ws7 = wb.create_sheet("关键词统计")
        ws7.append(["关键词", "出现次数", "最高热度", "平均热度", "最佳排名", "最新采集时间", "统计日期"])
        for row in keyword_stats:
            ws7.append([
                row.get("keyword", ""),
                row.get("appear_count", 0),
                row.get("max_hot_value", 0),
                row.get("avg_hot_value", 0),
                row.get("best_rank", ""),
                row.get("latest_fetch_time", ""),
                row.get("stat_date", ""),
            ])

        # Sheet8: 当前热搜
        ws8 = wb.create_sheet("当前热搜")
        ws8.append(["标题", "排名", "热度值", "来源", "采集时间"])
        for row in current_ranking:
            ws8.append([
                row.get("title", ""),
                row.get("rank_num", ""),
                row.get("hot_value", 0),
                row.get("source", ""),
                row.get("fetch_time", ""),
            ])

        # 保存文件
        filename = f"weibo_hot_enhanced_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        temp_path = Path(tempfile.gettempdir()) / filename
        wb.save(temp_path)

        return str(temp_path), filename

    except ImportError:
        raise Exception("openpyxl 未安装，请安装：pip install openpyxl")
    except Exception as e:
        raise Exception(f"导出增强报告 Excel 失败：{str(e)}")
